"""Excel export utilities for Physics QA data."""

from io import BytesIO
from typing import List, Dict, Any
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_jsonl_file(file_path: Path) -> List[Dict[str, Any]]:
    """
    Parse a JSONL file into a list of dictionaries.

    Args:
        file_path: Path to the JSONL file

    Returns:
        List of parsed JSON objects
    """
    items = []
    with open(file_path) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    items.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return items


def flatten_qa_item(item: Dict[str, Any], index: int) -> Dict[str, Any]:
    """
    Flatten a single QA item for Excel row.

    Args:
        item: The QA data point dictionary
        index: 0-based index of the item

    Returns:
        Flattened dictionary suitable for Excel row
    """
    rubric = item.get("rubric", {})
    final_answer = rubric.get("final_answer", {})

    return {
        "row_id": index + 1,
        "topic": item.get("topic", "Physics"),
        "query": item.get("query", ""),
        "response_answer": item.get("response_answer", ""),
        "response_reasoning": item.get("response_reasoning", ""),
        "rubric_total_points": rubric.get("total_points", 10),
        "rubric_final_answer_value": final_answer.get("value", ""),
        "rubric_final_answer_points": final_answer.get("points", 3),
        "rubric_final_answer_tolerance": final_answer.get("tolerance", ""),
        "rubric_common_errors": json.dumps(final_answer.get("common_errors", [])),
        "rubric_key_steps_count": len(rubric.get("key_steps", [])),
        "rubric_key_steps": json.dumps(rubric.get("key_steps", [])),
        "rubric_partial_credit_rules": json.dumps(rubric.get("partial_credit_rules", [])),
        "rubric_automatic_zero": json.dumps(rubric.get("automatic_zero", [])),
    }


def create_excel_workbook(items: List[Dict[str, Any]], filename: str) -> BytesIO:
    """
    Create Excel workbook from QA items.

    Args:
        items: List of QA data point dictionaries
        filename: Original filename (for metadata)

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Dataset"

    # Flatten items
    rows = [flatten_qa_item(item, i) for i, item in enumerate(items)]

    if not rows:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Write headers with styling
    headers = list(rows[0].keys())
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[header])

    # Auto-adjust column widths (with max limit for readability)
    for col in range(1, len(headers) + 1):
        max_length = 0
        for r in range(1, min(len(rows) + 2, 50)):  # Sample first 50 rows
            cell_value = ws.cell(row=r, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # Cap width: min 10, max 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
